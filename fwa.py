from collections import defaultdict
from datetime import timedelta
import json
import os
import random


import matplotlib.pyplot as plt
import numpy as np
from scipy.stats import norm
from scipy.spatial.distance import cdist
from tqdm import trange
import pandas as pd


class FWA:
    def __init__(self, func, dim, bounds, selection_method='distance', seed=None):
        self.func = func
        self.dim = dim
        self.bounds = bounds
        self.fireworks = []
        self.best_solution = None
        self.best_value = None
        self.history = []
        self.selection_method = selection_method
        self.n = None
        self.current_iter = 0

        self.__A_i = 0.0

        if seed is not None:
            np.random.seed(seed)
            random.seed(seed)

    def config(self, n=5, m=50, a=0.04, b=0.8, A_hat=40, m_hat=5, max_iter=100, J=0.2, J_hat=0.5):
        self.n = n
        self.m = m
        self.a = a
        self.b = b
        self.A_hat = A_hat
        self.m_hat = m_hat
        self.max_iter = max_iter
        self.J = J
        self.J_hat = J_hat
        self.__A_dynamic = self.A_hat

    def set_problem_context(self, start_date, n_days, n_employees, shift_id_to_index, shift_ids, shift_on_request, employee_id_to_index, cover_requirements):
        self.start_date = start_date
        self.n_days = n_days
        self.n_employees = n_employees
        self.shift_id_to_index = shift_id_to_index
        self.shift_ids = shift_ids
        self.shift_on_requests = shift_on_request
        self.employee_id_to_index = employee_id_to_index
        self.cover_requirements = cover_requirements

    
    def apply_joy_factor(self):
        J = self.J
        J_hat = self.J_hat

        if J == 0 or J_hat == 0:
            return

        if not hasattr(self, "start_date"):
            raise ValueError("Você precisa configurar o contexto do problema com set_problem_context().")

        n_mutate = max(1, int(self.n * J))
        mutate_indices = np.random.choice(len(self.fireworks), size=n_mutate, replace=False)

        # Pré-computar os índices dos dias que são sábado (5) ou domingo (6)
        weekend_days = [
            day for day in range(self.n_days)
            if (self.start_date + timedelta(days=day)).weekday() in [5, 6]
        ]

        # Índice do turno de folga
        off_index = self.shift_id_to_index.get("OFF", 0)

        for idx in mutate_indices:
            fw = self.fireworks[idx]
            schedule = np.rint(fw).astype(int).reshape((self.n_employees, self.n_days))

            # Gera máscara booleana aleatória para finais de semana
            mutation_mask = np.random.rand(self.n_employees, len(weekend_days)) < J_hat

            # Aplica OFF nos finais de semana conforme a máscara
            for i, day in enumerate(weekend_days):
                schedule[:, day][mutation_mask[:, i]] = off_index

            # Atualiza firework original
            self.fireworks[idx] = schedule.flatten().astype(float)


    def init_fireworks(self):
        self.fireworks = [
            np.random.uniform(low=[b[0] for b in self.bounds],
                              high=[b[1] for b in self.bounds])
            for _ in range(self.n)
        ]
        self.best_solution = min(self.fireworks, key=self.func)
        self.best_value = self.func(self.best_solution)
        self.history.append(self.best_value)

    def apply_shift_on_requests(self, schedule):
        """
        Aplica os ShiftOnRequests às escalas de forma a atender às preferências positivas dos funcionários.
        Cada entrada da lista shift_on_requests deve conter: 'EmployeeID', 'Date', 'ShiftTypeID', 'Weight'.
        """
        for req in self.shift_on_requests:
            emp_id = self.employee_id_to_index.get(req['EmployeeID'])
            day = (req['Date'] - self.start_date).days
            shift_symbol = req.get('ShiftTypeID')

            shift_idx = self.shift_id_to_index.get(shift_symbol)

            if (
                emp_id is not None and
                0 <= emp_id < self.n_employees and
                0 <= day < self.n_days and
                shift_idx is not None
            ):
                schedule[emp_id, day] = shift_idx

    def init_fireworks_smart(self, smart_ratio=0.7):
        assert hasattr(self, "n_employees") and hasattr(self, "n_days"), \
            "Use set_problem_context() antes de usar init_fireworks_smart()."

        self.fireworks = []

        off_index = self.shift_id_to_index.get("OFF", 4)
        available_shift_ids = [i for i in range(len(self.shift_ids)) if i != off_index]

        n_smart = int(self.n * smart_ratio)
        n_random = self.n - n_smart

        for _ in range(n_smart):
            schedule = np.zeros((self.n_employees, self.n_days), dtype=float)

            for emp in range(self.n_employees):
                # Atribui todos os dias com turnos aleatórios (sem OFF)
                schedule[emp] = np.random.choice(available_shift_ids, size=self.n_days)

                # Define 1 folga em sábado ou domingo
                weekend_days = [d for d in range(self.n_days)
                                if (self.start_date + timedelta(days=d)).weekday() in [5, 6]]
                if weekend_days:
                    off_day = random.choice(weekend_days)
                    schedule[emp, off_day] = off_index

                # Define +2 folgas aleatórias em dias que ainda não são OFF
                working_days = [d for d in range(self.n_days)
                                if schedule[emp, d] != off_index]
                extra_offs = random.sample(working_days, k=min(2, len(working_days)))
                for d in extra_offs:
                    schedule[emp, d] = off_index

            self.apply_shift_on_requests(schedule)
            self.fireworks.append(schedule.flatten())

        # Gera o restante da população de forma totalmente aleatória
        for _ in range(n_random):
            random_firework = np.random.uniform(
                low=[b[0] for b in self.bounds],
                high=[b[1] for b in self.bounds]
            )
            self.apply_shift_on_requests(schedule)
            self.fireworks.append(random_firework)

        self.best_solution = min(self.fireworks, key=self.func)
        self.best_value = self.func(self.best_solution)
        self.history.append(self.best_value)

    def shuffle_employee_subsequence(self, solution, max_block_size=7):
        """
        Para um empregado aleatório, embaralha um bloco consecutivo de turnos dentro do schedule.

        Args:
            solution: np.array (n_employees, n_days)
            max_block_size: tamanho máximo do bloco que pode ser embaralhado
        """
        emp = np.random.randint(0, self.n_employees)
        block_size = np.random.randint(2, min(max_block_size, self.n_days) + 1)
        start_day = np.random.randint(0, self.n_days - block_size + 1)
        end_day = start_day + block_size

        subseq = solution[emp, start_day:end_day]
        np.random.shuffle(subseq)
        solution[emp, start_day:end_day] = subseq

    def random_mutation(self, solution, mutation_rate=0.05):
        """
        Aplica mutação aleatória em alguns turnos da solução.

        Args:
            solution: np.array (n_employees, n_days)
            mutation_rate: probabilidade de mutar cada célula
        """
        for emp in range(self.n_employees):
            for day in range(self.n_days):
                if np.random.rand() < mutation_rate:
                    # Escolhe turno aleatório válido diferente do atual
                    current_shift = solution[emp, day]
                    possible_shifts = list(range(len(self.shift_ids)))
                    possible_shifts.remove(current_shift)

                    new_shift = np.random.choice(possible_shifts)
                    solution[emp, day] = new_shift

    def generate_initial_solution_cover(self, n_swaps=100):
        """
        Gera uma solução inicial tentando cobrir os turnos obrigatórios (cover_requirements)
        e depois aumenta diversidade trocando turnos entre empregados aleatoriamente.

        Args:
            n_swaps (int): número de trocas aleatórias a realizar para diversificar a solução.

        Retorna:
            np.array shape (n_employees, n_days) com índices dos turnos.
        """
        solution = np.full((self.n_employees, self.n_days), self.shift_id_to_index['OFF'], dtype=int)
        emp_order = np.random.permutation(self.n_employees)
        day_order = np.random.permutation(self.n_days)

        # Preencher para cobrir as demandas
        for day_offset in day_order:
            current_date = self.start_date + timedelta(days=int(day_offset))
            day_name = current_date.strftime("%A")

            for shift_id, required in self.cover_requirements.get(day_name, {}).items():
                shift_index = self.shift_id_to_index[shift_id]

                assigned = 0
                for emp in emp_order:
                    if solution[emp, day_offset] == self.shift_id_to_index['OFF']:
                        solution[emp, day_offset] = shift_index
                        assigned += 1
                        if assigned == required:
                            break

        # Aumenta a diversidade com trocas aleatórias entre empregados/dias
        for _ in range(n_swaps):
            emp1, emp2 = np.random.choice(self.n_employees, size=2, replace=False)
            day1, day2 = np.random.choice(self.n_days, size=2, replace=False)

            # Troca os turnos entre os dois empregados nos dias escolhidos
            temp = solution[emp1, day1]
            solution[emp1, day1] = solution[emp2, day2]
            solution[emp2, day2] = temp

        return solution

    
    def init_fireworks_smart_v2(self):
        self.fireworks = []
        for i in range(self.n):
            schedule = self.generate_initial_solution_cover()
            self.shuffle_employee_subsequence(schedule)
            self.random_mutation(schedule)
            self.apply_shift_on_requests(schedule)
            self.fireworks.append(schedule.flatten())
        
        self.best_solution = min(self.fireworks, key=self.func)
        self.best_value = self.func(self.best_solution)
        self.history.append(self.best_value)


    def run(self, verbose=False, log_freq=10):
        # self.init_fireworks()
        self.init_fireworks_smart(smart_ratio=0.5)
        # self.init_fireworks_smart_v2()

        self.current_iter = 0
        pbar = trange(self.max_iter, desc="FWA", dynamic_ncols=True)
        for i in pbar:
            self.iter()
            self.current_iter += 1
            pbar.set_description(f"FWA F={self.best_value:.0f} - A_i={self.__A_i:.8f} - A_dyn={self.__A_dynamic:.2f}")
            if verbose and (i % log_freq == 0 or i == self.max_iter - 1):
                print(f"Iter {i}: best = {self.best_value}")

    def get_dynamic_amplitude(self):
        initial = self.A_hat
        final = 0.2  # amplitude mínima no fim
        decay_rate = self.current_iter / (self.max_iter - 1 + 1e-8)
        return initial * (1 - decay_rate) + final * decay_rate


    def iter(self):
        sparks = []
        f_values = [self.func(fw) for fw in self.fireworks]
        ymax, ymin = max(f_values), min(f_values)
        total_diff = sum(ymax - f for f in f_values) + 1e-12

        amplitude_denom = sum(f - ymin for f in f_values) + 1e-12

        for i, fw in enumerate(self.fireworks):
            # cálculo da quant. de faíscas do fogo atual.
            # fogos com menor fitness geram mais faíscas
            s_i = self.m * (ymax - f_values[i] + 1e-12) / total_diff
            s_i = round(np.clip(s_i, self.a * self.m, self.b * self.m))

            # cálculo da amplitude da explosão. vai ser maior para fogos
            # com fitness ruins (vão explorar mais) e menor para fogos 
            # com fitness bons (busca será mais local)

            # A_i = self.A_hat * (f_values[i] - ymin + 1e-12) / amplitude_denom

            A_dynamic = self.get_dynamic_amplitude()
            A_i = A_dynamic * (f_values[i] - ymin + 1e-12) / amplitude_denom
            self.__A_i = A_i
            self.__A_dynamic = A_dynamic

            sparks += self.explode(fw, s_i, A_i)

        sparks += self.gaussian_explode()
        
        # sparks += self.shift_flip_explosion()

        self.fireworks = self.select(sparks + self.fireworks)
        best = min(self.fireworks, key=self.func)

        best_val = self.func(best)

        if self.best_value is None or best_val < self.best_value:
            self.best_solution = best
            self.best_value = best_val

        self.history.append(self.best_value)

        self.apply_joy_factor()


    def explode(self, fw, s_i, A_i):
        results = []
        for _ in range(s_i):
            spark = np.copy(fw)
            z = np.random.randint(1, self.dim + 1)
            idx = np.random.choice(self.dim, z, replace=False)
            for k in idx:
                h = A_i * np.random.uniform(-1, 1)
                spark[k] += h
                spark[k] = self.clip(spark[k], self.bounds[k])
            results.append(spark)
        return results

    def gaussian_explode(self):
        results = []
        for _ in range(self.m_hat):  # número de faíscas gaussianas

            # escolhe um firework existente aleatoriamente
            fw = random.choice(self.fireworks) 

            # cria uma cópia (para não alterar o original)
            spark = np.copy(fw)

            # número de dimensões a alterar (entre 1 e dim)
            z = np.random.randint(1, self.dim + 1)  

            # seleciona quais dimensões mudar
            idx = np.random.choice(self.dim, z, replace=False)  

            g = norm.rvs(loc=1, scale=1)  # G ~ N(1, 1)
            for k in idx:
                spark[k] *= g  # x' = x * G
                spark[k] = self.clip(spark[k], self.bounds[k])  # garante que está nos limites
            results.append(spark)  # adiciona a nova faísca
        return results
    

    def shift_flip_explosion(self):
        results = []
        for _ in range(self.m_hat):  # número de novas faíscas desse tipo
            fw = random.choice(self.fireworks)
            spark = np.copy(fw)
            z = np.random.randint(1, self.dim + 1)
            idx = np.random.choice(self.dim, z, replace=False)

            for k in idx:
                flip = 0.5 * random.choice([-2, -1, 0, 1, 2])
                spark[k] += flip
                spark[k] = self.clip(spark[k], self.bounds[k])

            results.append(spark)
        return results




    # def clip(self, val, bound):
    #     # Implementação reflexiva baseada no paper original (reflete no limite)
    #     min_b, max_b = bound
    #     while val < min_b or val > max_b:
    #         if val < min_b:
    #             val = min_b + (min_b - val)
    #         elif val > max_b:
    #             val = max_b - (val - max_b)
    #     # Correção final para garantir limite exato em casos de imprecisão numérica
    #     return np.clip(val, min_b, max_b)
    
    def clip(self, val, bound):
        min_b, max_b = bound
        range_b = max_b - min_b
        if val < min_b or val > max_b:
            val = min_b + abs((val - min_b) % (2 * range_b))
            if val > max_b:
                val = max_b - (val - max_b)
        return val


    def select(self, candidates):
        candidates = np.array(candidates)

        if self.selection_method == 'distance':
            return self.__select_distance_v2(candidates)
        elif self.selection_method == 'roulette':
            return self.__select_roulette(candidates)
        elif self.selection_method == 'tournament':
            return self.__select_tournament(candidates)
        else:
            return self.__select_random(candidates)
        
    def __select_distance(self, candidates):
        f_vals = [self.func(x) for x in candidates]

        # Elitismo: seleciona aleatoriamente um dos melhores em caso de empate
        best_val = np.min(f_vals)
        best_indices = [i for i, f in enumerate(f_vals) if f == best_val]
        best_idx = np.random.choice(best_indices)
        best = candidates[best_idx]  # elitismo

        # Cálculo eficiente das distâncias entre candidatos
        distances = cdist(candidates, candidates).sum(axis=1)
        probs = distances / distances.sum()

        # Remove o melhor selecionado da lista para evitar repetição
        candidates_wo_best = [candidates[i] for i in range(len(candidates)) if i != best_idx]
        probs = np.delete(probs, best_idx)

        selected = [best]  # sempre mantemos o melhor

        if len(candidates_wo_best) < self.n - 1:
            # Poucos candidatos viáveis: permitimos repetição com ruído
            chosen = np.random.choice(len(candidates_wo_best), self.n - 1, replace=True)
            for i in chosen:
                spark = np.copy(candidates_wo_best[i])
                noise = np.random.normal(loc=0, scale=0.2, size=spark.shape)  # ruído suave
                spark += noise
                spark = np.clip(spark, [b[0] for b in self.bounds], [b[1] for b in self.bounds])
                selected.append(spark)
        else:
            # Seleção por diversidade sem repetição
            probs /= probs.sum()
            chosen = np.random.choice(len(candidates_wo_best), self.n - 1, p=probs, replace=False)
            selected += [candidates_wo_best[i] for i in chosen]

        return selected

    def __select_distance_v2(self, candidates):
        f_vals = [self.func(x) for x in candidates]

        # Elitismo: seleciona aleatoriamente um dos melhores em caso de empate
        best_val = np.min(f_vals)
        best_indices = [i for i, f in enumerate(f_vals) if f == best_val]
        best_idx = np.random.choice(best_indices)
        best = candidates[best_idx]  # elitismo

        # Cálculo eficiente das distâncias entre candidatos
        distances = cdist(candidates, candidates).sum(axis=1)
        if distances.sum() == 0:
            probs = np.ones(len(distances)) / len(distances)
        else:
            probs = distances / distances.sum()

        # Remove o melhor selecionado da lista para evitar repetição
        candidates_wo_best = [candidates[i] for i in range(len(candidates)) if i != best_idx]
        probs = np.delete(probs, best_idx)

        # Normaliza as probabilidades de seleção
        if probs.sum() == 0:
            probs = np.ones_like(probs) / len(probs)
        else:
            probs /= probs.sum()

        selected = [best]  # sempre mantemos o melhor

        # Seleção por diversidade sem repetição
        chosen = np.random.choice(len(candidates_wo_best), self.n - 1, p=probs, replace=False)
        selected += [candidates_wo_best[i] for i in chosen]

        return selected

    
    def __select_roulette(self, candidates):
        f_vals = np.array([self.func(x) for x in candidates])
        best_idx = np.argmin(f_vals)
        best = candidates[best_idx]

        # Inverso da aptidão
        fitness_inv = np.max(f_vals) - f_vals + 1e-12

        # Zera a probabilidade do melhor (elitismo garantido fora da roleta)
        fitness_inv[best_idx] = 0.0

        # Renormaliza as probabilidades
        probs = fitness_inv / fitness_inv.sum()

        # Escolhe n - 1 sem repetição
        remaining_indices = np.random.choice(
            np.arange(len(candidates)),
            size=self.n - 1,
            replace=False,
            p=probs
        )

        selected = [best] + [candidates[i] for i in remaining_indices]
        return selected

    
    def __select_tournament(self, candidates):
        f_vals = [self.func(x) for x in candidates]
        best_idx = np.argmin(f_vals)
        best = candidates[best_idx]

        selected = [best]
        candidates_indices = list(range(len(candidates)))
        while len(selected) < self.n:
            i1, i2 = random.sample(candidates_indices, 2)
            winner = candidates[i1] if f_vals[i1] < f_vals[i2] else candidates[i2]
            if winner not in selected:
                selected.append(winner)
        return selected
    
    def __select_random(self, candidates):
        f_vals = [self.func(x) for x in candidates]
        best_idx = np.argmin(f_vals)
        best = candidates[best_idx]

        indices = np.arange(len(candidates))

        # evitar duplicar o melhor
        chosen = [i for i in np.random.choice(indices, self.n - 1, replace=False) if i != best_idx]
        while len(chosen) < self.n - 1:
            i = np.random.choice(indices)
            if i != best_idx and i not in chosen:
                chosen.append(i)

        selected = [best] + [candidates[i] for i in chosen]
        return selected
    
    def save_to_disc(self, path: str = "fwa_result.json"):
        # Salva a melhor solução, valor, histórico e parâmetros do problema em JSON.
        if self.best_solution is None:
            raise ValueError("Nenhuma solução foi encontrada ainda.")
        
        data = {
            "best_solution": self.best_solution.tolist(),
            "best_value": self.best_value,
            "history": self.history,
            "parameters": {
                "n": self.n,
                "m": self.m,
                "a": self.a,
                "b": self.b,
                "A_hat": self.A_hat,
                "m_hat": self.m_hat,
                "max_iter": self.max_iter,
                "dim": self.dim,
                "J" : self.J,
                "J_hat" : self.J_hat,
                # "bounds": [list(b) for b in self.bounds]
            }
        }

        # cria o diretório caso não exista
        dir_path = os.path.dirname(path)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)

        with open(path, "w") as f:
            json.dump(data, f, indent=4)
        
        print(f"Resultados salvos em: {path}")


    def load_best(self, path: str = "fwa_result.json"):
        # Carrega melhor solução, valor, histórico e parâmetros de um arquivo JSON.
        with open(path, "r") as f:
            data = json.load(f)

        self.best_solution = np.array(data["best_solution"])
        self.best_value = data["best_value"]
        self.history = data["history"]

        params = data.get("parameters", {})
        self.n = params.get("n", None)
        self.m = params.get("m", None)
        self.a = params.get("a", None)
        self.b = params.get("b", None)
        self.A_hat = params.get("A_hat", None)
        self.m_hat = params.get("m_hat", None)
        self.max_iter = params.get("max_iter", None)
        self.dim = params.get("dim", None)
        # self.bounds = [tuple(b) for b in params.get("bounds", [])]

        print(f"Solução carregada de {path} com valor: {self.best_value}")


    def plot_history_from_file(self, json_path, save_path=None):

        self.load_best(path=json_path)

        if not self.history:
            print("Histórico vazio.")
            return

        plt.figure(figsize=(12, 5))
        plt.plot(self.history, label=f"Fitness (best: {self.best_value})")
        plt.title("Evolução da Função Objetivo (FWA)")
        plt.xlabel("Iteração")
        plt.ylabel("Valor da função objetivo")
        plt.grid(True)
        plt.legend()
        plt.tight_layout()

        if save_path:
            plt.savefig(save_path, dpi=300)

        plt.show()
    


    def save_excel(self, filename="best_schedule.xlsx", shift_colors=None):
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        import calendar  # Para abreviações dos dias da semana

        if self.best_solution is None:
            raise ValueError("Nenhuma solução encontrada para salvar.")
        if not hasattr(self, "n_employees") or not hasattr(self, "n_days") or not hasattr(self, "shift_ids"):
            raise ValueError("Contexto do problema (n_employees, n_days, shift_ids) não configurado.")
        if not hasattr(self, "start_date"):
            raise ValueError("Atributo start_date não configurado na classe.")

        # Cores padrão baseadas no XML do dataset (cores convertidas para HEX)
        default_shift_colors = {
            "E": "FF0000",   # Red
            "D": "32CD32",   # Lime
            "L": "0000FF",   # Blue
            "N": "808080",   # Gray
            "OFF": "FFFFFF"  # Branco para OFF
        }

        if shift_colors is None:
            shift_colors = default_shift_colors

        # Converte vetor 1D em matriz (schedule)
        schedule = np.rint(self.best_solution).astype(int).reshape(self.n_employees, self.n_days)

        # Mapeia índices para labels dos turnos e converte "OFF" para "-"
        def map_shift_label(idx):
            label = self.shift_ids[idx]
            return "-" if label == "OFF" else label

        schedule_str = np.vectorize(map_shift_label)(schedule)

        # Cria DataFrame
        df = pd.DataFrame(schedule_str, 
                        index=[f"Employee {i+1}" for i in range(self.n_employees)],
                        columns=[f"Day {i+1}" for i in range(self.n_days)])

        # Salva dataframe sem estilo
        df.to_excel(filename, engine="openpyxl")

        # Abre arquivo para edição de estilos
        wb = load_workbook(filename)
        ws = wb.active

        # Inserir uma linha após o cabeçalho com abreviação do dia da semana
        # A linha 1 é o cabeçalho (Day 1, Day 2, ...), inserimos na linha 2 os dias da semana
        ws.insert_rows(2)

        for col_idx in range(2, 2 + self.n_days):
            current_date = self.start_date + timedelta(days=col_idx - 2)  # col_idx=2 corresponde a dia 0
            day_abbr = calendar.day_abbr[current_date.weekday()]  # Ex: Mon, Tue, Wed
            ws.cell(row=2, column=col_idx, value=day_abbr)

        # Aplicar cor nas células conforme shift_colors (dados começam na linha 3 agora)
        for row_idx in range(3, 3 + self.n_employees):
            for col_idx in range(2, 2 + self.n_days):
                cell = ws.cell(row=row_idx, column=col_idx)
                label = cell.value

                # OFF = "-" mantém fundo branco
                if label == "-":
                    continue

                color_hex = shift_colors.get(label, "FFFFFF")
                fill = PatternFill(start_color="FF" + color_hex, end_color="FF" + color_hex, fill_type="solid")
                cell.fill = fill

        wb.save(filename)
        print(f"Melhor escala salva em Excel colorido: {filename}")
